"""
Generate a fresh, current-state-of-the-code valuation-model spreadsheet.

Mirrors the form/layout of the original
`docs/fix-whatyourbusinesscouldbeworth/edgefour-valuation-model-v4.20.26_For Mike.xlsx`
but every value, multiplier, slider array, weight, bonus, range band,
methodology notice, and inline calculation scalar is parsed at runtime
from the live production source:

    api/_lib/calculator.ts
    public/industry-options-fragment.html

Each run writes a fresh, dated workbook. Hand it off to a non-technical
reviewer with confidence that it is a complete picture of what the app
actually does today — not a translation, not a snapshot, not patched.

Run:    python3 docs/valuation-model-spec/generate_valuation_spec.py
Output: docs/valuation-model-spec/edgefour-valuation-model-v{M}.{D}.{YY}.xlsx
        (e.g. edgefour-valuation-model-v4.24.26.xlsx)
"""

from __future__ import annotations

import os
import sys
from datetime import date

# Make the parser importable regardless of cwd.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from _parse_calculator import load as load_calculator_data


# ── LOAD ALL PRODUCTION DATA AT IMPORT ─────────────────────────────────────
_DATA                = load_calculator_data()
INDUSTRY_MULTIPLES   = _DATA.industry_multiples
OWNER_DEP_BY_SECTOR  = _DATA.owner_dep_by_sector
INDUSTRY_NOTICES     = _DATA.industry_notices
GROWTH_ADJ           = _DATA.growth_adj
RECURRING_ADJ        = _DATA.recurring_adj
CUST_CONC_ADJ        = _DATA.cust_conc_adj
SYSTEMS_ADJ          = _DATA.systems_adj
FIN_REC_ADJ          = _DATA.fin_rec_adj
INDUSTRY_GROUPS      = _DATA.industry_groups
NOTICE_KEYS          = set(INDUSTRY_NOTICES.keys())
RULES                = _DATA.rules
GIT_COMMIT           = _DATA.git_commit
SNAPSHOT_DATE        = _DATA.snapshot_date


# ── STYLES (matched to the original spec workbook's colour palette) ────────
NAVY        = "FF1B2A4A"
NAVY_MID    = "FF2D4A7A"
GOLD_PALE   = "FFF5EDD4"
ROW_TINT    = "FFF8F9FB"
WHITE       = "FFFFFFFF"
BLUE_TEXT   = "FF0000FF"
BLACK_TEXT  = "FF000000"
GRAY_TEXT   = "FF4A6080"

THIN_BORDER_SIDE = Side(style="thin", color="FFD0D7E2")
ALL_BORDERS = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE,
)

TITLE_FONT     = Font(name="Arial", size=14, bold=True, color=WHITE)
SUBTITLE_FONT  = Font(name="Arial", size=10, italic=True, color=BLACK_TEXT)
HEADER_FONT    = Font(name="Arial", size=10, bold=True, color=WHITE)
GROUP_FONT     = Font(name="Arial", size=10, bold=True, color=WHITE)
BODY_FONT      = Font(name="Arial", size=10, color=BLACK_TEXT)
INPUT_FONT     = Font(name="Calibri", size=11, color=BLUE_TEXT)
CALC_FONT      = Font(name="Calibri", size=11, color=BLACK_TEXT)
NOTE_FONT      = Font(name="Arial", size=9, italic=True, color=GRAY_TEXT)
BOLD_FONT      = Font(name="Arial", size=10, bold=True, color=BLACK_TEXT)

TITLE_FILL     = PatternFill("solid", fgColor=NAVY)
HEADER_FILL    = PatternFill("solid", fgColor=NAVY)
GROUP_FILL     = PatternFill("solid", fgColor=NAVY_MID)
SECTION_FILL   = PatternFill("solid", fgColor=GOLD_PALE)


def style_title(ws, cell_range: str, text: str, *, height: int = 28) -> None:
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = text
    top_left.font = TITLE_FONT
    top_left.fill = TITLE_FILL
    top_left.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
    row_idx = int("".join(c for c in cell_range.split(":")[0] if c.isdigit()))
    ws.row_dimensions[row_idx].height = height


def style_subtitle(ws, cell_range: str, text: str, *, height: int = 60) -> None:
    ws.merge_cells(cell_range)
    top_left = ws[cell_range.split(":")[0]]
    top_left.value = text
    top_left.font = SUBTITLE_FONT
    top_left.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
    row_idx = int("".join(c for c in cell_range.split(":")[0] if c.isdigit()))
    ws.row_dimensions[row_idx].height = height


def style_header_row(ws, row_idx: int, headers: list[str]) -> None:
    for col_idx, value in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = ALL_BORDERS
    ws.row_dimensions[row_idx].height = 30


def style_group_row(ws, row_idx: int, label: str, span: int) -> None:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span)
    cell = ws.cell(row=row_idx, column=1, value=label)
    cell.font = GROUP_FONT
    cell.fill = GROUP_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    cell.border = ALL_BORDERS
    ws.row_dimensions[row_idx].height = 22


def write_input(cell, value) -> None:
    cell.value = value
    cell.font = INPUT_FONT
    cell.alignment = Alignment(vertical="center", horizontal="right")
    cell.border = ALL_BORDERS


def write_body(cell, value, *, bold: bool = False) -> None:
    cell.value = value
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True, indent=1)
    cell.border = ALL_BORDERS


def write_calc(cell, value) -> None:
    cell.value = value
    cell.font = CALC_FONT
    cell.alignment = Alignment(vertical="center", horizontal="right")
    cell.border = ALL_BORDERS


# ── README ─────────────────────────────────────────────────────────────────

def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 110

    style_title(ws, "A1:B1",
        "EdgeFour Value — Current Production Model "
        f"(snapshot {SNAPSHOT_DATE} @ git {GIT_COMMIT})")
    style_subtitle(ws, "A2:B2",
        "Every value in this workbook was parsed at generation time from "
        "api/_lib/calculator.ts and public/industry-options-fragment.html, so it is a "
        "complete and current picture of what the app does today. Re-run "
        "generate_valuation_spec.py at any time to regenerate against the latest code. "
        "Blue cells = inputs the model treats as editable. Black cells = derived/calculated.",
        height=70,
    )

    rows = [
        ("Source of truth",                f"api/_lib/calculator.ts ({len(INDUSTRY_MULTIPLES)} industries, 6 sliders, bonuses, scoring, trajectory)"),
        ("Industry labels",                f"public/industry-options-fragment.html ({sum(len(items) for _, items in INDUSTRY_GROUPS)} entries across {len(INDUSTRY_GROUPS)} categories)"),
        ("Sheet: Industry Base Multiples", "Baseline EBITDA multiple assigned to each industry before any slider or bonus adjustment."),
        ("Sheet: Owner Dependency Adj",    "Industry-specific multiple adjustments for the Owner Dependency / Management Team slider — the only factor with industry-specific sensitivity."),
        ("Sheet: Universal Factor Adj",    "Multiple adjustments for revenue growth, recurring revenue, customer concentration, systems maturity, and financial records — identical across all industries."),
        ("Sheet: Bonus Adjustments",       "Years-in-business and revenue-scale bonuses — flat absolute multiple bumps added on top of slider adjustments."),
        ("Sheet: Score Weights",           "Weights used to compute the 0–100 Business Value Score, plus the score-band thresholds."),
        ("Sheet: Valuation Range Bands",   f"Multipliers used to spread the base valuation into a low / base / high range (current production: {RULES.range_low:g} / 1.00 / {RULES.range_high:g})."),
        ("Sheet: Trajectory Logic",        f"How the “After Key Improvements” trajectory uplift is computed (clampDelta capped at {RULES.trajectory_max_levels} levels per factor; total capped at baseMultiple + {RULES.trajectory_total_cap_above_base:g})."),
        ("Sheet: Methodology Notices",     "Industries flagged with a non-EBITDA primary methodology — full notice text shown to the user on the results page."),
        ("Sheet: Full Calc Example",       "End-to-end worked example combining every component for a sample HVAC business. Edit the blue cells to model different scenarios."),
    ]
    for r_idx, (col_a, col_b) in enumerate(rows, start=4):
        cell_a = ws.cell(row=r_idx, column=1, value=col_a)
        cell_b = ws.cell(row=r_idx, column=2, value=col_b)
        cell_a.font = BOLD_FONT
        cell_b.font = BODY_FONT
        cell_a.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
        cell_b.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
        ws.row_dimensions[r_idx].height = 32


# ── INDUSTRY BASE MULTIPLES ────────────────────────────────────────────────

def build_industry_multiples(wb: Workbook) -> None:
    ws = wb.create_sheet("Industry Base Multiples")
    for col, width in {"A": 30, "B": 48, "C": 14, "D": 16, "E": 28}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:E1",
        f"Industry Base EBITDA Multiples — All {len(INDUSTRY_MULTIPLES)} Industries")
    style_subtitle(ws, "A2:E2",
        "Base Multiple is the EBITDA multiple for a typical well-run business in this "
        "category before any slider or bonus adjustments. Industries marked with ⚠ use a "
        "non-EBITDA primary methodology — see the Methodology Notices sheet for the full "
        "user-facing explanation. Blue values are the inputs the calculator reads directly."
    )
    style_header_row(ws, 3, ["Category", "Industry / Sub-Type", "Base Multiple", "Method Note", "Industry Code (slug)"])

    row = 4
    for category, items in INDUSTRY_GROUPS:
        style_group_row(ws, row, category, span=5)
        row += 1
        for slug, label in items:
            multiple = INDUSTRY_MULTIPLES.get(slug)
            method_note = "⚠ See notice" if slug in NOTICE_KEYS else ""
            write_body(ws.cell(row=row, column=1), category)
            write_body(ws.cell(row=row, column=2), label)
            write_input(ws.cell(row=row, column=3), multiple)
            ws.cell(row=row, column=3).number_format = "0.00\"x\""
            write_body(ws.cell(row=row, column=4), method_note)
            write_body(ws.cell(row=row, column=5), slug)
            ws.cell(row=row, column=5).font = NOTE_FONT
            row += 1

    ws.freeze_panes = "A4"


# ── OWNER DEPENDENCY ADJUSTMENTS ───────────────────────────────────────────

def build_owner_dep(wb: Workbook) -> None:
    ws = wb.create_sheet("Owner Dependency Adj")
    for col, width in {"A": 30, "B": 48, "C": 12, "D": 12, "E": 14, "F": 12, "G": 14, "H": 14, "I": 24}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:I1",
        "Owner Dependency / Management Team — Industry-Specific Multiple Adjustments")
    style_subtitle(ws, "A2:I2",
        "Owner dependency is the only factor with industry-specific adjustments. "
        "L1 = fully owner-dependent. L3 = neutral baseline. L5 = the business runs without "
        "the owner. Industries where client relationships are personal (professional services, "
        "healthcare) penalise high owner dependency much more heavily than systemisable "
        "industries (retail, e-commerce). All values feed directly into the live calculation.",
        height=70,
    )
    style_header_row(ws, 3, [
        "Category", "Industry",
        "L1\n(Fully Dep.)", "L2\n(High)", "L3\n(Moderate)",
        "L4\n(Low)", "L5\n(Minimal)", "Max Swing\n(L1→L5)",
        "Industry Code (slug)",
    ])

    row = 4
    for category, items in INDUSTRY_GROUPS:
        style_group_row(ws, row, category, span=9)
        row += 1
        for slug, label in items:
            arr = OWNER_DEP_BY_SECTOR.get(slug, OWNER_DEP_BY_SECTOR["other"])
            swing = arr[4] - arr[0]
            write_body(ws.cell(row=row, column=1), category)
            write_body(ws.cell(row=row, column=2), label)
            for offset, value in enumerate(arr):
                cell = ws.cell(row=row, column=3 + offset)
                write_input(cell, value)
                cell.number_format = "+0.00;-0.00;0.00"
            swing_cell = ws.cell(row=row, column=8)
            write_calc(swing_cell, swing)
            swing_cell.number_format = "+0.00\"x\";-0.00\"x\";0.00\"x\""
            write_body(ws.cell(row=row, column=9), slug)
            ws.cell(row=row, column=9).font = NOTE_FONT
            row += 1

    ws.freeze_panes = "A4"


# ── UNIVERSAL FACTOR ADJUSTMENTS ───────────────────────────────────────────

def build_universal(wb: Workbook) -> None:
    ws = wb.create_sheet("Universal Factor Adj")
    for col, width in {"A": 38, "B": 12, "C": 12, "D": 14, "E": 12, "F": 14, "G": 14}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:G1", "Universal Multiple Adjustments — All Industries")
    style_subtitle(ws, "A2:G2",
        "These adjustments apply equally across every industry. L3 is the neutral baseline "
        "(no adjustment) for most sliders. Customer Concentration is INVERTED — L1 "
        "(diversified) is best, L5 (single-customer) is worst. Values are added to the "
        "industry base multiple before the years and revenue-scale bonuses.",
        height=60,
    )
    style_header_row(ws, 3, [
        "Factor", "L1 (Worst)", "L2", "L3 (Neutral)", "L4", "L5 (Best)", "Max Swing",
    ])

    factors = [
        ("Revenue Growth (3-Year Trend)",                          GROWTH_ADJ,    "L1=Declining, L3=Stable (2–5%), L5=Hypergrowth (15%+)"),
        ("Recurring Revenue %",                                    RECURRING_ADJ, "L1=<10%, L3=25–50%, L5=75%+"),
        ("Largest Customer Concentration (INVERTED)",              CUST_CONC_ADJ, "L1=<5% spread (best), L5=>50% (worst). Note negative values at L4–L5."),
        ("Systems & Process Maturity",                             SYSTEMS_ADJ,   "L1=nothing documented, L3=partial, L5=full SOPs + software"),
        ("Financial Record & Operational Data Quality",            FIN_REC_ADJ,   "L1=messy/incomplete, L3=adequate, L5=reviewed/audited"),
    ]

    row = 4
    for label, arr, sub in factors:
        write_body(ws.cell(row=row, column=1), label, bold=True)
        for offset, value in enumerate(arr):
            cell = ws.cell(row=row, column=2 + offset)
            write_input(cell, value)
            cell.number_format = "+0.00;-0.00;0.00"
        swing = arr[4] - arr[0]
        swing_cell = ws.cell(row=row, column=7)
        write_calc(swing_cell, swing)
        swing_cell.number_format = "+0.00\"x\";-0.00\"x\";0.00\"x\""
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=7)
        sub_cell = ws.cell(row=row + 1, column=1, value=sub)
        sub_cell.font = NOTE_FONT
        sub_cell.alignment = Alignment(vertical="center", horizontal="left", indent=2)
        row += 2

    ws.freeze_panes = "A4"


# ── BONUS ADJUSTMENTS ──────────────────────────────────────────────────────

def _fmt_money(n: int) -> str:
    return f"${n:,}"


def build_bonus_adjustments(wb: Workbook) -> None:
    ws = wb.create_sheet("Bonus Adjustments")
    for col, width in {"A": 32, "B": 28, "C": 18, "D": 60}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:D1", "Bonus Adjustments — Years in Business & Revenue Scale")
    style_subtitle(ws, "A2:D2",
        "Both bonuses are flat absolute multiple bumps added on top of the slider "
        f"adjustments. The final multiple is then floored at {RULES.final_multiple_floor:g}x.",
    )
    style_header_row(ws, 3, ["Bonus", "Threshold", "Multiple Bump", "Notes"])

    rows = [
        ("Years in Business", f"≥ {RULES.years_bonus_high_threshold} years",
            RULES.years_bonus_high, "Demonstrates durability through full economic cycles."),
        ("Years in Business",
            f"≥ {RULES.years_bonus_low_threshold} years (and < {RULES.years_bonus_high_threshold})",
            RULES.years_bonus_low, "Survival past the start-up risk window."),
        ("Years in Business", f"< {RULES.years_bonus_low_threshold} years",
            0.00, "No bonus."),
        ("Revenue Scale", f"≥ {_fmt_money(RULES.revenue_bonus_high_threshold)}",
            RULES.revenue_bonus_high, "Scaled, lower-middle-market profile attracts a wider buyer pool."),
        ("Revenue Scale",
            f"≥ {_fmt_money(RULES.revenue_bonus_low_threshold)} (and < {_fmt_money(RULES.revenue_bonus_high_threshold)})",
            RULES.revenue_bonus_low, "Past the lower-middle-market threshold."),
        ("Revenue Scale", f"< {_fmt_money(RULES.revenue_bonus_low_threshold)}",
            0.00, "No bonus."),
    ]
    for r_idx, (factor, threshold, bump, note) in enumerate(rows, start=4):
        write_body(ws.cell(row=r_idx, column=1), factor, bold=True)
        write_body(ws.cell(row=r_idx, column=2), threshold)
        bump_cell = ws.cell(row=r_idx, column=3)
        write_input(bump_cell, bump)
        bump_cell.number_format = "+0.00\"x\";-0.00\"x\";0.00\"x\""
        write_body(ws.cell(row=r_idx, column=4), note)


# ── SCORE WEIGHTS ──────────────────────────────────────────────────────────

def build_score_weights(wb: Workbook) -> None:
    ws = wb.create_sheet("Score Weights")
    for col, width in {"A": 36, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:G1", "Business Value Score — Factor Weights (0–100 Scale)")
    style_subtitle(ws, "A2:G2",
        f"Score starts at {RULES.score_base} (all sliders at L3 = neutral). Each factor "
        f"adds or subtracts (adjustment × weight). +{RULES.score_years_bonus_pts} bonus "
        f"points if years in business ≥ {RULES.score_years_bonus_threshold}. Final score is "
        f"rounded and clamped to [{RULES.score_min}, {RULES.score_max}]. Owner dependency "
        "uses the industry-specific array from the Owner Dependency Adj sheet.",
        height=60,
    )
    style_header_row(ws, 3, ["Factor", "Weight", "L1 pts", "L2 pts", "L3 pts", "L4 pts", "L5 pts"])

    universal_factors = [
        ("Revenue Growth (growth)",                          RULES.score_weight_growth,    GROWTH_ADJ),
        ("Recurring Revenue % (recurring)",                  RULES.score_weight_recurring, RECURRING_ADJ),
        ("Customer Concentration (custConc) — INVERTED",     RULES.score_weight_cust_conc, CUST_CONC_ADJ),
        ("Systems & Process Maturity (systems)",             RULES.score_weight_systems,   SYSTEMS_ADJ),
        ("Financial Record Quality (finRec)",                RULES.score_weight_fin_rec,   FIN_REC_ADJ),
    ]

    row = 4
    for label, weight, arr in universal_factors:
        write_body(ws.cell(row=row, column=1), label, bold=True)
        write_calc(ws.cell(row=row, column=2), weight)
        for offset, value in enumerate(arr):
            cell = ws.cell(row=row, column=3 + offset)
            write_calc(cell, round(value * weight, 2))
            cell.number_format = "+0.0;-0.0;0.0"
        row += 1

    write_body(ws.cell(row=row, column=1), "Owner Dep / Mgmt Team (ownerDep)", bold=True)
    write_calc(ws.cell(row=row, column=2), RULES.score_weight_owner_dep)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    note_cell = ws.cell(row=row, column=3,
        value=f"Industry-specific array × {RULES.score_weight_owner_dep} (see Owner Dependency Adj sheet)")
    note_cell.font = NOTE_FONT
    note_cell.alignment = Alignment(vertical="center", horizontal="center")
    note_cell.border = ALL_BORDERS
    row += 1

    write_body(ws.cell(row=row, column=1),
        f"Years in Business (≥ {RULES.score_years_bonus_threshold} years)", bold=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    note_cell = ws.cell(row=row, column=2,
        value=f"+{RULES.score_years_bonus_pts} pts if years ≥ {RULES.score_years_bonus_threshold} (no bonus otherwise)")
    note_cell.font = NOTE_FONT
    note_cell.alignment = Alignment(vertical="center", horizontal="center")
    note_cell.border = ALL_BORDERS
    row += 2

    style_header_row(ws, row, ["Score Range", "Label", "Descriptor Shown on Results Page", "", "", "", ""])
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    row += 1

    bands = [
        ("80–100", "Excellent",  "Excellent — your business is well-positioned for a premium exit."),
        ("65–79",  "Good",       "Good — solid fundamentals with specific opportunities to increase value."),
        ("50–64",  "Fair",       "Fair — meaningful improvements could significantly lift your valuation."),
        (f"{RULES.score_min}–49", "Developing", "Developing — there are clear steps that could substantially increase what buyers will pay."),
    ]
    for r in bands:
        write_body(ws.cell(row=row, column=1), r[0])
        write_body(ws.cell(row=row, column=2), r[1], bold=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        write_body(ws.cell(row=row, column=3), r[2])
        row += 1


# ── VALUATION RANGE BANDS ──────────────────────────────────────────────────

def build_range_bands(wb: Workbook) -> None:
    ws = wb.create_sheet("Valuation Range Bands")
    for col, width in {"A": 28, "B": 14, "C": 28, "D": 80}.items():
        ws.column_dimensions[col].width = width

    band_pct = (RULES.range_high - 1.0) * 100
    style_title(ws, "A1:D1", "Valuation Range Bands")
    style_subtitle(ws, "A2:D2",
        "The base valuation (Adjusted EBITDA × Final Multiple) is spread into a range "
        f"using these multipliers. Production currently uses {RULES.range_low:g} / 1.00 / "
        f"{RULES.range_high:g} — a ±{band_pct:g}% band around the central estimate, "
        "calibrated to typical buyer-negotiation outcomes.",
    )
    style_header_row(ws, 3, ["Scenario", "Multiplier", "Formula", "Rationale"])

    low_pct = (1.0 - RULES.range_low) * 100
    high_pct = (RULES.range_high - 1.0) * 100
    rows = [
        ("Conservative Value", RULES.range_low,
            f"baseVal × {RULES.range_low:g}",
            f"Reflects buyer negotiation, deal risk, and the reality that most businesses do not achieve full asking price. ~{low_pct:g}% haircut."),
        ("Base Estimated Value", 1.00,
            "adjEBITDA × finalMultiple",
            "The central estimate the tool stands behind as the most likely fair-market value through a well-run sale process."),
        ("Optimistic Value", RULES.range_high,
            f"baseVal × {RULES.range_high:g}",
            f"Reflects a competitive process with multiple buyers, a strategic acquirer paying for synergies, or a top-of-market position. ~{high_pct:g}% premium."),
    ]
    for r_idx, (label, mult, formula, rationale) in enumerate(rows, start=4):
        write_body(ws.cell(row=r_idx, column=1), label, bold=True)
        m_cell = ws.cell(row=r_idx, column=2)
        write_input(m_cell, mult)
        m_cell.number_format = "0.00"
        write_body(ws.cell(row=r_idx, column=3), formula)
        write_body(ws.cell(row=r_idx, column=4), rationale)
        ws.row_dimensions[r_idx].height = 36


# ── TRAJECTORY LOGIC ───────────────────────────────────────────────────────

def build_trajectory(wb: Workbook) -> None:
    ws = wb.create_sheet("Trajectory Logic")
    for col, width in {"A": 36, "B": 60, "C": 70}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:C1", "Trajectory Projection — “After Key Improvements”")
    style_subtitle(ws, "A2:C2",
        "Shows what the business could be worth after improving its weakest factors. "
        f"Each factor improvement is capped at {RULES.trajectory_max_levels} slider levels. "
        f"Total multiple gain is capped at baseMultiple + {RULES.trajectory_total_cap_above_base:g}. "
        "The trajectory uplift is then displayed on the results page as “$X increase in value with key improvements.”",
        height=55,
    )
    style_header_row(ws, 3, ["Step", "Formula / Logic", "Notes"])

    rows = [
        ("1. clampDelta(arr, idx, max=2)",
         "adj[min(idx+max,4)] − adj[idx-1]",
         f"Looks at the actual adjustment array values — never linearly interpolated. Caps target at L5 (max={RULES.trajectory_max_levels})."),
        ("2. Per-factor delta",
         "delta = clampDelta(adjArray, currentLevel, 2)",
         "Computed independently for growth, recurring, custConc (using reversed array), systems, finRec, and ownerDep."),
        ("3. Filter & rank",
         "Keep factors whose current level ≤ 3 (custConc ≥ 3) AND delta > 0; sort by delta DESC; take top 2.",
         "These two factors drive the displayed “top opportunities” on the results card."),
        ("4. trajMultipleGain",
         "Sum of the deltas of the top-2 selected factors",
         "Total potential multiple improvement from acting on the two biggest gaps."),
        ("5. trajMultiple",
         f"MAX(finalMultiple, MIN(finalMultiple + gain, baseMultiple + {RULES.trajectory_total_cap_above_base:g}))",
         f"Floor: never drops below today's final multiple. Cap: never exceeds baseMultiple + {RULES.trajectory_total_cap_above_base:g}."),
        ("6. trajVal",
         "adjEBITDA × trajMultiple",
         "Trajectory valuation (the “after key improvements” base value)."),
        ("7. trajUplift",
         "trajVal − baseVal",
         "Dollar improvement displayed on the results page."),
        ("8. New range",
         f"low = origLow + trajUplift × {RULES.range_low:g}; high = origHigh + trajUplift × {RULES.range_high:g}",
         "The trajectory range shown on the results page (also rounded to whole dollars)."),
    ]
    row = 4
    for step, formula, note in rows:
        write_body(ws.cell(row=row, column=1), step, bold=True)
        write_body(ws.cell(row=row, column=2), formula)
        write_body(ws.cell(row=row, column=3), note)
        ws.row_dimensions[row].height = 38
        row += 1


# ── METHODOLOGY NOTICES ────────────────────────────────────────────────────

def build_methodology_notices(wb: Workbook) -> None:
    ws = wb.create_sheet("Methodology Notices")
    for col, width in {"A": 26, "B": 60, "C": 110}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:C1",
        "Methodology Notices — Industries Where EBITDA Is Not the Primary Method")
    style_subtitle(ws, "A2:C2",
        "These industries are flagged in the tool with an inline notice on the results page "
        "explaining why an EBITDA multiple is directional only. The full notice text below is "
        "exactly what the user sees.",
        height=50,
    )
    style_header_row(ws, 3, ["Industry Code", "Title", "Body"])

    label_lookup = {slug: label for _, items in INDUSTRY_GROUPS for slug, label in items}

    row = 4
    for slug, payload in INDUSTRY_NOTICES.items():
        write_body(ws.cell(row=row, column=1), f"{label_lookup.get(slug, slug)}\n({slug})")
        write_body(ws.cell(row=row, column=2), payload["title"], bold=True)
        write_body(ws.cell(row=row, column=3), payload["body"])
        ws.row_dimensions[row].height = 110
        row += 1


# ── FULL CALC EXAMPLE (live formulas matching production code) ────────────

def build_full_example(wb: Workbook) -> None:
    ws = wb.create_sheet("Full Calc Example")
    for col, width in {"A": 38, "B": 22, "C": 22, "D": 60}.items():
        ws.column_dimensions[col].width = width

    style_title(ws, "A1:D1", "End-to-End Calculation Example (live formulas)")
    style_subtitle(ws, "A2:D2",
        "Worked example for an HVAC business. Blue cells = inputs you can edit — every "
        "downstream value updates automatically. Black cells are calculated. The formulas "
        "exactly mirror api/_lib/calculator.ts: comp adj → adjusted EBITDA → slider sum → "
        "years bonus → revenue bonus → final multiple → valuation range.",
        height=55,
    )
    style_header_row(ws, 3, ["Step", "Input / Value", "Result", "Notes"])

    def section(row: int, label: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=label)
        c.font = BOLD_FONT
        c.fill = SECTION_FILL
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)

    R_INDUSTRY        = 5
    R_BASE_MULT       = 6
    R_YEARS           = 7
    R_REVENUE         = 8
    R_EBITDA          = 9
    R_OWNER_SAL       = 10
    R_MARKET_SAL      = 11
    R_ADDBACKS        = 12
    R_COMP_ADJ        = 15
    R_ADJ_EBITDA      = 16
    R_GROWTH          = 19
    R_OWNER_DEP       = 20
    R_RECURRING       = 21
    R_CUST_CONC       = 22
    R_SYSTEMS         = 23
    R_FIN_REC         = 24
    R_OWNER_DEP_ADJ   = 26
    R_GROWTH_ADJ_VAL  = 27
    R_REC_ADJ_VAL     = 28
    R_CUST_ADJ_VAL    = 29
    R_SYS_ADJ_VAL     = 30
    R_FIN_ADJ_VAL     = 31
    R_SLIDER_SUM      = 32
    R_YEARS_BONUS     = 33
    R_REV_BONUS       = 34
    R_FINAL_MULT      = 35
    R_BASE_VAL        = 38
    R_LOW_VAL         = 39
    R_HIGH_VAL        = 40
    R_SCORE           = 43

    # ── INPUTS ────────────────────────────────────────────────────────────
    section(4, "INPUTS")
    write_body(ws.cell(row=R_INDUSTRY, column=1),    "Industry (display name)")
    write_input(ws.cell(row=R_INDUSTRY, column=2),   "HVAC")
    write_body(ws.cell(row=R_INDUSTRY, column=4),    "Edit the base multiple cell below if you change industry.")

    write_body(ws.cell(row=R_BASE_MULT, column=1),   "Industry Base Multiple")
    write_input(ws.cell(row=R_BASE_MULT, column=2),  INDUSTRY_MULTIPLES["hvac"])
    ws.cell(row=R_BASE_MULT, column=2).number_format = "0.00\"x\""
    write_body(ws.cell(row=R_BASE_MULT, column=4),   "Pulled from the Industry Base Multiples sheet.")

    write_body(ws.cell(row=R_YEARS, column=1),       "Years in Business")
    write_input(ws.cell(row=R_YEARS, column=2),      14)
    write_body(ws.cell(row=R_YEARS, column=4),
        f"Triggers +{RULES.years_bonus_high:.2f}x bonus when ≥ {RULES.years_bonus_high_threshold}; "
        f"+{RULES.years_bonus_low:.2f}x when ≥ {RULES.years_bonus_low_threshold}.")

    write_body(ws.cell(row=R_REVENUE, column=1),     "Annual Revenue")
    write_input(ws.cell(row=R_REVENUE, column=2),    4_500_000)
    ws.cell(row=R_REVENUE, column=2).number_format = "$#,##0"
    write_body(ws.cell(row=R_REVENUE, column=4),
        f"Triggers +{RULES.revenue_bonus_low:.2f}x bonus ≥ {_fmt_money(RULES.revenue_bonus_low_threshold)}; "
        f"+{RULES.revenue_bonus_high:.2f}x ≥ {_fmt_money(RULES.revenue_bonus_high_threshold)}.")

    write_body(ws.cell(row=R_EBITDA, column=1),      "Reported EBITDA")
    write_input(ws.cell(row=R_EBITDA, column=2),     550_000)
    ws.cell(row=R_EBITDA, column=2).number_format = "$#,##0"

    write_body(ws.cell(row=R_OWNER_SAL, column=1),   "Owner Total Compensation")
    write_input(ws.cell(row=R_OWNER_SAL, column=2),  220_000)
    ws.cell(row=R_OWNER_SAL, column=2).number_format = "$#,##0"
    write_body(ws.cell(row=R_OWNER_SAL, column=4),   "W-2 + distributions running through P&L.")

    write_body(ws.cell(row=R_MARKET_SAL, column=1),  "Market Rate for Owner Role")
    write_input(ws.cell(row=R_MARKET_SAL, column=2), 140_000)
    ws.cell(row=R_MARKET_SAL, column=2).number_format = "$#,##0"
    write_body(ws.cell(row=R_MARKET_SAL, column=4),  "Replacement GM/CEO cost.")

    write_body(ws.cell(row=R_ADDBACKS, column=1),    "Other Add-Backs")
    write_input(ws.cell(row=R_ADDBACKS, column=2),   25_000)
    ws.cell(row=R_ADDBACKS, column=2).number_format = "$#,##0"
    write_body(ws.cell(row=R_ADDBACKS, column=4),    "Non-recurring legal expense (or any other add-back).")

    # ── ADJUSTED EBITDA ───────────────────────────────────────────────────
    section(14, "ADJUSTED EBITDA")
    write_body(ws.cell(row=R_COMP_ADJ, column=1), "Compensation Adjustment")
    write_calc(ws.cell(row=R_COMP_ADJ, column=3), f"=B{R_OWNER_SAL}-B{R_MARKET_SAL}")
    ws.cell(row=R_COMP_ADJ, column=3).number_format = "$#,##0"
    write_body(ws.cell(row=R_COMP_ADJ, column=4), "Owner Total Comp − Market Rate (can go negative).")

    write_body(ws.cell(row=R_ADJ_EBITDA, column=1), "Adjusted EBITDA", bold=True)
    write_calc(ws.cell(row=R_ADJ_EBITDA, column=3), f"=B{R_EBITDA}+C{R_COMP_ADJ}+B{R_ADDBACKS}")
    ws.cell(row=R_ADJ_EBITDA, column=3).number_format = "$#,##0"
    write_body(ws.cell(row=R_ADJ_EBITDA, column=4), "= EBITDA + Comp Adj + Add-Backs. Primary valuation driver.")

    # ── SLIDERS ───────────────────────────────────────────────────────────
    section(18, "SLIDER INPUTS (1–5)")
    sliders = [
        (R_GROWTH,    "Revenue Growth (growth)",                      4),
        (R_OWNER_DEP, "Owner Dep / Mgmt Team (ownerDep)",             3),
        (R_RECURRING, "Recurring Revenue (recurring)",                2),
        (R_CUST_CONC, "Customer Concentration (custConc)",            2),
        (R_SYSTEMS,   "Systems & Process (systems)",                  3),
        (R_FIN_REC,   "Financial Records (finRec)",                   4),
    ]
    for r, label, val in sliders:
        write_body(ws.cell(row=r, column=1), label)
        write_input(ws.cell(row=r, column=2), val)
        write_body(ws.cell(row=r, column=4), "Whole integer 1–5. L3 = neutral baseline.")

    # ── MULTIPLE ──────────────────────────────────────────────────────────
    section(25, "MULTIPLE CALCULATION")

    write_body(ws.cell(row=R_OWNER_DEP_ADJ, column=1), "Owner Dep adj (HVAC array INDEX)")
    write_input(ws.cell(row=R_OWNER_DEP_ADJ, column=2),
                ", ".join(f"{v:+.2f}" for v in OWNER_DEP_BY_SECTOR["hvac"]))
    write_calc(ws.cell(row=R_OWNER_DEP_ADJ, column=3),
               f"=INDEX({{{','.join(str(v) for v in OWNER_DEP_BY_SECTOR['hvac'])}}},B{R_OWNER_DEP})")
    ws.cell(row=R_OWNER_DEP_ADJ, column=3).number_format = "+0.00;-0.00;0.00"
    write_body(ws.cell(row=R_OWNER_DEP_ADJ, column=4), "Industry-specific. Replace the array if the industry changes.")

    universal_rows = [
        (R_GROWTH_ADJ_VAL, "Growth adj",            GROWTH_ADJ,    R_GROWTH),
        (R_REC_ADJ_VAL,    "Recurring adj",         RECURRING_ADJ, R_RECURRING),
        (R_CUST_ADJ_VAL,   "Customer Conc adj",     CUST_CONC_ADJ, R_CUST_CONC),
        (R_SYS_ADJ_VAL,    "Systems adj",           SYSTEMS_ADJ,   R_SYSTEMS),
        (R_FIN_ADJ_VAL,    "Financial Records adj", FIN_REC_ADJ,   R_FIN_REC),
    ]
    for r, label, arr, slider_row in universal_rows:
        write_body(ws.cell(row=r, column=1), label)
        write_input(ws.cell(row=r, column=2),
                    ", ".join(f"{v:+.2f}" for v in arr))
        write_calc(ws.cell(row=r, column=3),
                   f"=INDEX({{{','.join(str(v) for v in arr)}}},B{slider_row})")
        ws.cell(row=r, column=3).number_format = "+0.00;-0.00;0.00"
        write_body(ws.cell(row=r, column=4), "Universal across all industries.")

    write_body(ws.cell(row=R_SLIDER_SUM, column=1), "Slider Adjustments Sum")
    write_calc(ws.cell(row=R_SLIDER_SUM, column=3),
               f"=C{R_OWNER_DEP_ADJ}+C{R_GROWTH_ADJ_VAL}+C{R_REC_ADJ_VAL}+C{R_CUST_ADJ_VAL}+C{R_SYS_ADJ_VAL}+C{R_FIN_ADJ_VAL}")
    ws.cell(row=R_SLIDER_SUM, column=3).number_format = "+0.00\"x\";-0.00\"x\";0.00\"x\""
    write_body(ws.cell(row=R_SLIDER_SUM, column=4), "Sum of all six slider adjustments.")

    write_body(ws.cell(row=R_YEARS_BONUS, column=1), "Years Bonus")
    write_calc(ws.cell(row=R_YEARS_BONUS, column=3),
               f"=IF(B{R_YEARS}>={RULES.years_bonus_high_threshold},{RULES.years_bonus_high},"
               f"IF(B{R_YEARS}>={RULES.years_bonus_low_threshold},{RULES.years_bonus_low},0))")
    ws.cell(row=R_YEARS_BONUS, column=3).number_format = "+0.00\"x\";-0.00\"x\";0.00\"x\""
    write_body(ws.cell(row=R_YEARS_BONUS, column=4),
        f"+{RULES.years_bonus_high:.2f}x ≥ {RULES.years_bonus_high_threshold} yrs; "
        f"+{RULES.years_bonus_low:.2f}x ≥ {RULES.years_bonus_low_threshold} yrs; else 0.")

    write_body(ws.cell(row=R_REV_BONUS, column=1), "Revenue Scale Bonus")
    write_calc(ws.cell(row=R_REV_BONUS, column=3),
               f"=IF(B{R_REVENUE}>={RULES.revenue_bonus_high_threshold},{RULES.revenue_bonus_high},"
               f"IF(B{R_REVENUE}>={RULES.revenue_bonus_low_threshold},{RULES.revenue_bonus_low},0))")
    ws.cell(row=R_REV_BONUS, column=3).number_format = "+0.00\"x\";-0.00\"x\";0.00\"x\""
    write_body(ws.cell(row=R_REV_BONUS, column=4),
        f"+{RULES.revenue_bonus_high:.2f}x ≥ {_fmt_money(RULES.revenue_bonus_high_threshold)}; "
        f"+{RULES.revenue_bonus_low:.2f}x ≥ {_fmt_money(RULES.revenue_bonus_low_threshold)}; else 0.")

    write_body(ws.cell(row=R_FINAL_MULT, column=1), "Final Multiple", bold=True)
    write_calc(ws.cell(row=R_FINAL_MULT, column=3),
               f"=MAX({RULES.final_multiple_floor},B{R_BASE_MULT}+C{R_SLIDER_SUM}+C{R_YEARS_BONUS}+C{R_REV_BONUS})")
    ws.cell(row=R_FINAL_MULT, column=3).number_format = "0.00\"x\""
    write_body(ws.cell(row=R_FINAL_MULT, column=4),
        f"= MAX({RULES.final_multiple_floor:g}, base + slider sum + years bonus + revenue bonus).")

    # ── RANGE ─────────────────────────────────────────────────────────────
    section(37, "VALUATION RANGE")
    write_body(ws.cell(row=R_BASE_VAL, column=1), "Base Estimated Value", bold=True)
    write_calc(ws.cell(row=R_BASE_VAL, column=3), f"=C{R_ADJ_EBITDA}*C{R_FINAL_MULT}")
    ws.cell(row=R_BASE_VAL, column=3).number_format = "$#,##0"
    write_body(ws.cell(row=R_BASE_VAL, column=4), "= Adjusted EBITDA × Final Multiple.")

    write_body(ws.cell(row=R_LOW_VAL, column=1), "Conservative Value")
    write_calc(ws.cell(row=R_LOW_VAL, column=3), f"=C{R_BASE_VAL}*{RULES.range_low}")
    ws.cell(row=R_LOW_VAL, column=3).number_format = "$#,##0"
    write_body(ws.cell(row=R_LOW_VAL, column=4), f"Base × {RULES.range_low:g}.")

    write_body(ws.cell(row=R_HIGH_VAL, column=1), "Optimistic Value")
    write_calc(ws.cell(row=R_HIGH_VAL, column=3), f"=C{R_BASE_VAL}*{RULES.range_high}")
    ws.cell(row=R_HIGH_VAL, column=3).number_format = "$#,##0"
    write_body(ws.cell(row=R_HIGH_VAL, column=4), f"Base × {RULES.range_high:g}.")

    # ── SCORE ─────────────────────────────────────────────────────────────
    section(42, "BUSINESS VALUE SCORE")
    write_body(ws.cell(row=R_SCORE, column=1), "Value Score (0–100)", bold=True)
    write_calc(
        ws.cell(row=R_SCORE, column=3),
        f"=MIN({RULES.score_max},MAX({RULES.score_min},ROUND({RULES.score_base}"
        f"+C{R_GROWTH_ADJ_VAL}*{RULES.score_weight_growth}"
        f"+C{R_OWNER_DEP_ADJ}*{RULES.score_weight_owner_dep}"
        f"+C{R_REC_ADJ_VAL}*{RULES.score_weight_recurring}"
        f"+C{R_CUST_ADJ_VAL}*{RULES.score_weight_cust_conc}"
        f"+C{R_SYS_ADJ_VAL}*{RULES.score_weight_systems}"
        f"+C{R_FIN_ADJ_VAL}*{RULES.score_weight_fin_rec}"
        f"+IF(B{R_YEARS}>={RULES.score_years_bonus_threshold},{RULES.score_years_bonus_pts},0)"
        ",0)))"
    )
    ws.cell(row=R_SCORE, column=3).number_format = "0"
    write_body(ws.cell(row=R_SCORE, column=4),
        f"= ROUND({RULES.score_base} + Σ(adj × weight) + ({RULES.score_years_bonus_pts} if years ≥ "
        f"{RULES.score_years_bonus_threshold}), 0), clamped to [{RULES.score_min}, {RULES.score_max}].")


# ── BUILD ENTRY POINT ──────────────────────────────────────────────────────

def _output_filename() -> str:
    today = date.today()
    return f"edgefour-valuation-model-v{today.month}.{today.day}.{today.year % 100}.xlsx"


def main() -> None:
    wb = Workbook()
    build_readme(wb)
    build_industry_multiples(wb)
    build_owner_dep(wb)
    build_universal(wb)
    build_bonus_adjustments(wb)
    build_score_weights(wb)
    build_range_bands(wb)
    build_trajectory(wb)
    build_methodology_notices(wb)
    build_full_example(wb)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, _output_filename())
    wb.save(out_path)

    print(f"Wrote: {out_path}")
    print(f"  Snapshot:        {SNAPSHOT_DATE} @ git {GIT_COMMIT}")
    print(f"  Industries:      {len(INDUSTRY_MULTIPLES)}")
    print(f"  Owner-dep:       {len(OWNER_DEP_BY_SECTOR)} arrays")
    print(f"  Notices:         {len(INDUSTRY_NOTICES)}")
    print(f"  Range bands:     {RULES.range_low:g} / 1.00 / {RULES.range_high:g}")
    print(f"  Years bonus:     +{RULES.years_bonus_low:g}x ≥ {RULES.years_bonus_low_threshold}yr / "
          f"+{RULES.years_bonus_high:g}x ≥ {RULES.years_bonus_high_threshold}yr")
    print(f"  Revenue bonus:   +{RULES.revenue_bonus_low:g}x ≥ {_fmt_money(RULES.revenue_bonus_low_threshold)} / "
          f"+{RULES.revenue_bonus_high:g}x ≥ {_fmt_money(RULES.revenue_bonus_high_threshold)}")


if __name__ == "__main__":
    main()
